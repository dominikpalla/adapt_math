"""
Detekce cross-source duplicit v celé DB (cv/umat/e/ss/olz atd.).

Strategie:
  1) Exact match: hash normalizovaného content_latex + expected.
  2) Fuzzy match: SequenceMatcher ratio ≥ 0.85 mezi úlohami z RŮZNÝCH
     zdrojů (jiný task_id prefix).

Zdroje se odvozují z prefixu task_id (znaky do prvního číslice):
  cv, umat, e, ss, olz1, olz2, olzk, olte, atd.

Výstup XLSX + terminal report.

Použití:
    DATABASE_URL=... python scripts/find_cross_source_duplicates.py \\
        --out-xlsx /tmp/cross_duplicates.xlsx
"""
from __future__ import annotations
import argparse, os, re, sys, json
from pathlib import Path
from collections import defaultdict
from difflib import SequenceMatcher
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from database import init_db
from model import MathTask


def normalize_latex(s: str) -> str:
    if not s:
        return ""
    s = s.strip()
    # Sjednotit synonyma
    replacements = [
        (r'\$', ''),  # obal $
        (r'\\limits', ''),  # \lim vs \lim\limits
        (r'\\dfrac', r'\\frac'),
        (r'\\tg\b', r'\\tan'),
        (r'\\cotg\b', r'\\cot'),
        (r'\\arctg\b', r'\\arctan'),
        (r'\\arccotg\b', r'\\arccot'),
        (r'\\Rightarrow', r'\\Ra'),
        (r'\\left\s*\(', r'('),
        (r'\\right\s*\)', r')'),
        (r'\\left\s*\[', r'['),
        (r'\\right\s*\]', r']'),
        (r'\\left\s*\\\{', r'\\{'),
        (r'\\right\s*\\\}', r'\\}'),
        (r'\\langle', r'['),
        (r'\\rangle', r']'),
    ]
    for old, new in replacements:
        s = re.sub(old, new, s)
    s = re.sub(r'\s+', '', s)  # AGRESIVNÍ: úplně odstranit whitespace
    return s.lower()


def source_prefix(task_id: str) -> str:
    """Vrátí prefix jako 'cv', 'umat', 'e', 'ss', 'olz1', 'olz2', ...
    definuje ,,zdroj'' úlohy."""
    m = re.match(r'([a-z]+\d*)_', task_id)
    if not m:
        # e.g. 'olz1_lim_001' -> 'olz1' (odchytí olz1/olz2/olzk/olte)
        return task_id.split('_')[0] if '_' in task_id else task_id
    return m.group(1)


def content_key(task) -> str:
    parts = [normalize_latex(task.content_latex or "")]
    if isinstance(task.results, list):
        for r in task.results:
            if not isinstance(r, dict): continue
            exp = r.get("expected")
            if isinstance(exp, (str, int, float)):
                parts.append(normalize_latex(str(exp)))
            elif isinstance(exp, list):
                for e in exp:
                    parts.append(normalize_latex(str(e)))
    return "||".join(parts)


def content_for_fuzzy(task) -> str:
    return normalize_latex(task.content_latex or "")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--db-url", default=os.environ.get("DATABASE_URL"))
    parser.add_argument("--out-xlsx", default="/tmp/cross_duplicates.xlsx")
    parser.add_argument("--fuzzy-threshold", type=float, default=0.85)
    args = parser.parse_args()
    if not args.db_url:
        print("Chyba: DATABASE_URL.", file=sys.stderr); return 2

    SessionLocal = init_db(args.db_url)
    sess = SessionLocal()
    try:
        tasks = sess.query(MathTask).all()
        print(f"Načteno {len(tasks)} úloh celkem.")

        # 1) Cross-source exact groups
        by_key = defaultdict(list)
        for t in tasks:
            k = content_key(t)
            if k.strip("|"):
                by_key[k].append(t)
        # Jen skupiny obsahující >=2 různé zdroje
        cross_exact = []
        for g in by_key.values():
            if len(g) < 2: continue
            sources = {source_prefix(t.task_id) for t in g}
            if len(sources) >= 2:
                cross_exact.append(g)
        print(f"\nCROSS-SOURCE EXACT match skupin: {len(cross_exact)}")
        for g in cross_exact[:30]:
            ids = [t.task_id for t in g]
            print(f"  {ids}")
            print(f"    {g[0].content_latex[:100]}")

        # 2) Cross-source fuzzy match
        # Filtrujeme jen úlohy s content ≥ 15 chars, přeskočíme exact-cross duplicates.
        exact_ids = {t.task_id for g in cross_exact for t in g}
        candidates = [t for t in tasks if t.task_id not in exact_ids and len(content_for_fuzzy(t)) >= 15]
        print(f"\nFuzzy porovnání mezi {len(candidates)} úlohami (bez exact-cross)...")

        # Optimalizace: bucketovat po délce content (±30%) a source prefix
        fuzzy_pairs = []
        # Přesně: pro každou úlohu porovnávej jen s ostatními jiného zdroje s délkou v pásmu
        norm_map = [(t, content_for_fuzzy(t), source_prefix(t.task_id)) for t in candidates]
        n = len(norm_map)
        for i in range(n):
            ta, ai, srca = norm_map[i]
            for j in range(i+1, n):
                tb, aj, srcb = norm_map[j]
                if srca == srcb: continue  # cross-source only
                lr = min(len(ai), len(aj)) / max(len(ai), len(aj))
                if lr < 0.7: continue
                sim = SequenceMatcher(None, ai, aj).ratio()
                if sim >= args.fuzzy_threshold:
                    fuzzy_pairs.append((ta, tb, sim))

        fuzzy_pairs.sort(key=lambda x: -x[2])
        print(f"\nCROSS-SOURCE FUZZY párů (sim >= {args.fuzzy_threshold}): {len(fuzzy_pairs)}")
        for a, b, sim in fuzzy_pairs[:30]:
            print(f"  {sim:.2f}: {a.task_id} <-> {b.task_id}")

        # ---- XLSX out ----
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            print("openpyxl chybí, XLSX přeskočen."); return 0

        wb = Workbook()
        ws = wb.active
        ws.title = "Cross-source EXACT"
        ws.append(["Skupina", "task_id", "zdroj", "content_latex", "expected"])
        for c in ws[1]:
            c.font = Font(bold=True); c.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        for idx, g in enumerate(cross_exact, 1):
            for t in g:
                exp = ""
                if isinstance(t.results, list) and t.results and isinstance(t.results[0], dict):
                    exp = str(t.results[0].get("expected", ""))[:200]
                ws.append([idx, t.task_id, source_prefix(t.task_id),
                           (t.content_latex or "")[:400], exp])
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 60
        ws.column_dimensions['E'].width = 30

        ws2 = wb.create_sheet("Cross-source FUZZY")
        ws2.append(["Sim", "A: task_id", "A: zdroj", "A: content", "B: task_id", "B: zdroj", "B: content"])
        for c in ws2[1]:
            c.font = Font(bold=True); c.fill = PatternFill(start_color="FFF3CC", end_color="FFF3CC", fill_type="solid")
        for a, b, sim in fuzzy_pairs:
            ws2.append([
                round(sim, 3),
                a.task_id, source_prefix(a.task_id), (a.content_latex or "")[:400],
                b.task_id, source_prefix(b.task_id), (b.content_latex or "")[:400],
            ])
        ws2.column_dimensions['A'].width = 7
        ws2.column_dimensions['B'].width = 14
        ws2.column_dimensions['C'].width = 8
        ws2.column_dimensions['D'].width = 50
        ws2.column_dimensions['E'].width = 14
        ws2.column_dimensions['F'].width = 8
        ws2.column_dimensions['G'].width = 50
        for row in ws2.iter_rows(min_row=2):
            for c in row:
                c.alignment = Alignment(vertical="top", wrap_text=True)

        wb.save(args.out_xlsx)
        print(f"\nXLSX: {args.out_xlsx}")
        return 0
    finally:
        sess.close()


if __name__ == "__main__":
    sys.exit(main())
