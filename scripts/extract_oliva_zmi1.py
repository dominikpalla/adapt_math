"""
Extrakce úloh z Blackboard QTI-ish exportu (Oliva ZMI1 fondy).

Vstup: složka s rozbalenými fondy (každý fond je adresář s res*.dat XML soubory).
Výstup: JSON preview (jeden řádek per úloha) + XLSX summary pro tým.

Klíčová věc: v `mat_formattedtext type="HTML"` je math zapsaný jako MathML,
ALE s vloženou `<annotation encoding="LaTeX">HOTOVÝ_LATEX</annotation>`.
Extrakce = najít všechny annotation bloky a nahradit MathML kus za `$LATEX$`.

Typy úloh:
  - Numeric               → AdaptMath decimal (přesná odpověď v <varequal>)
  - Multiple Choice       → AdaptMath multiple_choice (jedna správná)
  - True/False            → AdaptMath multiple_choice (2 volby)
  - Fill in the Blank Plus → AdaptMath multiple_choice (ruční review)
  - File Upload           → SKIP (nepoužitelné v adaptivním systému)
  - Jumbled Sentence      → SKIP (převádět ručně)
  - Short Response        → SKIP (ruční review)
  - Multiple Answer       → AdaptMath multiple_choice s více správnými
  - Ordering              → SKIP (edge case)

Použití:
    python scripts/extract_oliva_zmi1.py --src <extracted_dir> \
        --out-json <preview.jsonl> --out-xlsx <summary.xlsx>

Nikdy nezapisuje do DB.
"""
from __future__ import annotations

import argparse
import html
import json
import os
import re
import sys
import unicodedata
from pathlib import Path
from typing import Any

# ------ HTML / MathML → text s $LaTeX$ ---------------------------------

_ANNOT_RE = re.compile(
    r'<math[^>]*>.*?<annotation[^>]*encoding="LaTeX"[^>]*>(.*?)</annotation>.*?</math>',
    re.DOTALL,
)


def html_to_text_with_latex(raw: str) -> str:
    """Vezme raw HTML s vnořenými MathML+annotation a vrátí čistý text
    s math částmi obalenými do $...$."""
    if not raw:
        return ""
    # 1) unescape HTML entity dvakrát (Blackboard je encoduje 2×)
    s = html.unescape(html.unescape(raw))
    # 2) nahradit celý <math>...</math> blok za $LATEX$
    def _sub(m):
        latex = m.group(1).strip()
        # znormalizovat whitespace v LaTeXu
        latex = re.sub(r'\s+', ' ', latex).strip()
        return f'${latex}$'
    s = _ANNOT_RE.sub(_sub, s)
    # 3) strip zbylých HTML tagů (typicky <p>, <br />, <em>, atd.)
    s = re.sub(r'<[^>]+>', ' ', s)
    # 4) čistka whitespace
    s = re.sub(r'\s+', ' ', s).strip()
    # 5) unicode NFC
    s = unicodedata.normalize('NFC', s)
    return s


# ------ Parser jednotlivé úlohy -----------------------------------------

_MAT_TEXT_RE = re.compile(
    r'<mat_formattedtext[^>]*>(.*?)</mat_formattedtext>',
    re.DOTALL,
)
_QTYPE_RE = re.compile(r'<bbmd_questiontype>([^<]+)</bbmd_questiontype>')

# Extract v <presentation> blok
_PRESENTATION_RE = re.compile(
    r'<presentation.*?>(.*?)</presentation>',
    re.DOTALL,
)
_RESPROC_RE = re.compile(
    r'<resprocessing.*?>(.*?)</resprocessing>',
    re.DOTALL,
)
_ITEMFEEDBACK_RE = re.compile(
    r'<itemfeedback[^>]*>(.*?)</itemfeedback>',
    re.DOTALL,
)

# MC options v <response_label ident="X"> ... </response_label>
_RESPONSE_LABEL_RE = re.compile(
    r'<response_label\s+ident="([^"]+)"[^>]*>(.*?)</response_label>',
    re.DOTALL,
)

# Correct answer v <respcondition>
_VAREQUAL_RE = re.compile(
    r'<varequal[^>]*>([^<]+)</varequal>',
)
_SETVAR_RE = re.compile(
    r'<setvar[^>]*action="Set"[^>]*>SCORE\.max</setvar>',
    re.IGNORECASE,
)

# Split kompletního XML podle <item ...>
_ITEM_SPLIT_RE = re.compile(r'(?=<item\s[^>]*>)')


def extract_items_from_xml(xml_content: str) -> list[str]:
    """Vrátí list surových XML kousků, jeden per <item>."""
    # Najdeme startovní pozice všech <item ...>
    matches = list(re.finditer(r'<item\s[^>]*>', xml_content))
    if not matches:
        return []
    starts = [m.start() for m in matches]
    ends = starts[1:] + [len(xml_content)]
    result = []
    for s, e in zip(starts, ends):
        chunk = xml_content[s:e]
        # Ořízni na </item>
        idx = chunk.find('</item>')
        if idx > 0:
            chunk = chunk[:idx + len('</item>')]
        result.append(chunk)
    return result


def parse_item(item_xml: str) -> dict[str, Any]:
    """Zparsuje jednu úlohu na dict s klíči:
       question_type, content_latex, options (list dict), correct_key,
       correct_value (pro Numeric), raw_correct_answers (list stringů)."""
    result: dict[str, Any] = {
        "question_type": None,
        "content_latex": "",
        "options": [],
        "correct_key": None,
        "correct_value": None,
        "raw_correct_answers": [],
    }
    # 1) question type
    m = _QTYPE_RE.search(item_xml)
    if m:
        result["question_type"] = m.group(1).strip()

    # 2) presentation → question text (první mat_formattedtext v presentation)
    p = _PRESENTATION_RE.search(item_xml)
    if p:
        # Najdi VŠECHNY mat_formattedtext v presentation (může být více: text +
        # sub-labely). První = zadání.
        parts = _MAT_TEXT_RE.findall(p.group(1))
        if parts:
            texts = [html_to_text_with_latex(x) for x in parts if x.strip()]
            result["content_latex"] = " ".join(texts)

        # 3) MC options (v presentation blok pod render_choice)
        for lm in _RESPONSE_LABEL_RE.finditer(p.group(1)):
            key = lm.group(1)
            body = lm.group(2)
            # options mají mat_formattedtext s labelem
            lbl_parts = _MAT_TEXT_RE.findall(body)
            label = " ".join(html_to_text_with_latex(x) for x in lbl_parts if x.strip())
            if label:
                result["options"].append({"key": key, "label_latex": label})

    # 4) correct answer z resprocessing
    r = _RESPROC_RE.search(item_xml)
    if r:
        resp_body = r.group(1)
        # Pro každý <respcondition>: hledáme buď <setvar SCORE.max>
        # (klasické MC/T/F) NEBO <displayfeedback linkrefid="correct"> (Numeric).
        for rc in re.finditer(
            r'<respcondition[^>]*>(.*?)</respcondition>', resp_body, re.DOTALL
        ):
            body = rc.group(1)
            is_correct = (
                _SETVAR_RE.search(body) or
                re.search(r'<displayfeedback\s+linkrefid="correct"', body)
            )
            if is_correct:
                # extract varequal, varlte, vargte
                for veq in _VAREQUAL_RE.finditer(body):
                    val = html.unescape(veq.group(1)).strip()
                    if val:
                        result["raw_correct_answers"].append(val)
                # Numeric range (pro toleranci)
                gte = re.search(r'<vargte[^>]*>([^<]+)</vargte>', body)
                lte = re.search(r'<varlte[^>]*>([^<]+)</varlte>', body)
                if gte and lte:
                    try:
                        g = float(gte.group(1)); l = float(lte.group(1))
                        result["_num_range"] = (g, l)
                    except (TypeError, ValueError):
                        pass

    # 5) pro Numeric: correct_value je jediné číslo v raw_correct_answers
    if result["question_type"] == "Numeric" and result["raw_correct_answers"]:
        result["correct_value"] = result["raw_correct_answers"][0]

    # 6) pro MC/True-False: correct_key = ten label ident, který je v raw_correct_answers
    if result["question_type"] in ("Multiple Choice", "True/False", "Multiple Answer"):
        opt_keys = {o["key"] for o in result["options"]}
        matching = [v for v in result["raw_correct_answers"] if v in opt_keys]
        if matching:
            result["correct_key"] = matching[0] if len(matching) == 1 else matching

    # 7) True/False: v Blackboardu jsou options "true"/"false", často bez
    #    viditelných labelů. Vždycky přeplácnout kompletní pair Pravda/Nepravda.
    if result["question_type"] == "True/False":
        cs_map = {"true": "Pravda", "false": "Nepravda"}
        result["options"] = [
            {"key": "true", "label_latex": "Pravda"},
            {"key": "false", "label_latex": "Nepravda"},
        ]
        # correct_key z raw_correct_answers
        for v in result["raw_correct_answers"]:
            vl = v.lower()
            if vl in ("true", "false"):
                result["correct_key"] = vl
                break

    return result


# ------ Mapping do AdaptMath schematu -----------------------------------

# Fond → kategorie AdaptMath (viz TASK_CATEGORIES). None = ruční review.
FOND_TO_CATEGORY = {
    "Z1 - inverzní funkce":                   "Funkce - Inverzní funkce",
    "Z1 - limita":                             "Limita - VOAL (dosazení)",
    "Z1 - negace výroku":                     "Výroková logika",
    "Z1 - slovní úloha":                       None,
    "Z1 - sudálichá":                          "Funkce - Sudá/lichá",
    "Z2  - asymptoty":                         "Funkce - Asymptota",
    "Z2  - tečna a normála":                   "Funkce - Tečna ke grafu",
    "Z2 - definice derivace":                  "Derivace - Sčítání",
    "Z2 - derivace":                           "Derivace - Sčítání",
    "Z2 - druhá derivace":                     "Derivace - Vyšší řády",
    "Z2 - extrémy na intervalu":               "Monotonie a extrémy",
    "Z2 - lHospitalovo pravidlo":              "Limita - Lhopitalovo pravidlo",
    "Z2 - lokální extrémy":                    "Monotonie a extrémy",
    "ZK - Taylorův polynom":                   "Derivace - Taylorův polynom",
    "ZK - aproximace pomocí diferenciálu":     "Derivace - Diferenciál",
    "ZK - integrál":                           "Primitivní funkce - Sčítání",
    "ZK - průběh funkce":                      "Průběh funkce",
    "ZK - slovní úloha":                       None,
    "ZK - tečna a normála":                    "Funkce - Tečna ke grafu",
    "ZK-TEORIE - derivace":                    "Derivace - Sčítání",
    "ZK-TEORIE - limita a spojitost":          "Spojitost - Spojitost",
    "ZK-TEORIE - primitivní funkce":           "Primitivní funkce - Sčítání",
    "ZK-TEORIE - realná funkce":               "Funkce - Definiční obor",
}

# Fond → task_id prefix
FOND_TO_PREFIX = {
    "Z1 - inverzní funkce":                    "olz1_inv",
    "Z1 - limita":                              "olz1_lim",
    "Z1 - negace výroku":                      "olz1_neg",
    "Z1 - slovní úloha":                        "olz1_slo",
    "Z1 - sudálichá":                           "olz1_sud",
    "Z2  - asymptoty":                          "olz2_asy",
    "Z2  - tečna a normála":                    "olz2_tec",
    "Z2 - definice derivace":                   "olz2_def",
    "Z2 - derivace":                            "olz2_der",
    "Z2 - druhá derivace":                      "olz2_der2",
    "Z2 - extrémy na intervalu":                "olz2_ext",
    "Z2 - lHospitalovo pravidlo":               "olz2_lho",
    "Z2 - lokální extrémy":                     "olz2_lok",
    "ZK - Taylorův polynom":                    "olzk_tay",
    "ZK - aproximace pomocí diferenciálu":      "olzk_apr",
    "ZK - integrál":                            "olzk_int",
    "ZK - průběh funkce":                       "olzk_prb",
    "ZK - slovní úloha":                        "olzk_slo",
    "ZK - tečna a normála":                     "olzk_tec",
    "ZK-TEORIE - derivace":                     "olte_der",
    "ZK-TEORIE - limita a spojitost":           "olte_lim",
    "ZK-TEORIE - primitivní funkce":            "olte_pf",
    "ZK-TEORIE - realná funkce":                "olte_rf",
}


def _strip_fond_name(fond: str) -> str:
    """Zbavit fond names prefixu 'Pool_ExportFile_KIKM-ZMI1_' a .zip."""
    n = fond
    prefix = "Pool_ExportFile_KIKM-ZMI1_"
    if n.startswith(prefix):
        n = n[len(prefix):]
    if n.endswith(".zip"):
        n = n[:-4]
    return n


def to_adaptmath(item: dict[str, Any], fond: str, idx: int) -> dict[str, Any] | None:
    """Konverze parsed item na AdaptMath task dict, nebo None pokud skip."""
    qtype = item["question_type"]

    # Skip nepoužitelné typy
    if qtype in ("File Upload", "Jumbled Sentence", "Ordering", "Short Response"):
        return None

    short_fond = _strip_fond_name(fond)
    prefix = FOND_TO_PREFIX.get(short_fond, "oliva")
    task_id = f"{prefix}_{idx:03d}"

    category = FOND_TO_CATEGORY.get(short_fond)

    result: dict[str, Any] = {
        "task_id": task_id,
        "content_latex": item["content_latex"],
        "category": category,
        "cognitive_load": "C",
        "irt_difficulty": 0,
        "irt_discrimination": 1.0,
        "_source_fond": short_fond,
        "_source_qtype": qtype,
        "_review_needed": False,
        "_review_reason": "",
    }

    # results podle typu
    if qtype == "Numeric":
        val = item.get("correct_value")
        num_range = item.get("_num_range")
        tolerance = 0.01
        if num_range:
            g, l = num_range
            tolerance = round((l - g) / 2, 6)
        try:
            result["results"] = [{
                "key": "vysledek",
                "label_latex": "= ",
                "type": "decimal",
                "expected": float(val),
                "tolerance": tolerance,
            }]
        except (TypeError, ValueError):
            result["results"] = [{
                "key": "vysledek",
                "label_latex": "= ",
                "type": "mathlive",
                "expected": val or "",
                "tolerance": 0,
            }]
            result["_review_needed"] = True
            result["_review_reason"] = "Numeric bez validního čísla"
    elif qtype in ("Multiple Choice", "True/False"):
        opts = item.get("options", [])
        ck = item.get("correct_key")
        if not opts or not ck:
            return None  # rozbité
        # Blackboard používá UUID jako keys. True/False má already true/false.
        # Pro MC přemapovat na a, b, c, d podle pořadí.
        if qtype == "Multiple Choice":
            key_map = {o["key"]: chr(ord('a') + i) for i, o in enumerate(opts[:26])}
            opts = [{"key": key_map[o["key"]], "label_latex": o["label_latex"]}
                    for o in opts if o["key"] in key_map]
            if isinstance(ck, list):
                ck = [key_map.get(k, k) for k in ck]
            else:
                ck = key_map.get(ck, ck)
        result["results"] = [{
            "key": "vysledek",
            "label_latex": r"\text{Odpověď: }",
            "type": "multiple_choice",
            "options": opts,
            "expected": ck,
        }]
    elif qtype == "Multiple Answer":
        opts = item.get("options", [])
        ck = item.get("correct_key")
        if not opts or not ck:
            return None
        # AdaptMath má multi tak, že expected je list klíčů (nový, viz task-checker)
        result["results"] = [{
            "key": "vysledek",
            "label_latex": r"\text{Odpovědi (více správných): }",
            "type": "multiple_choice",
            "options": opts,
            "expected": ck if isinstance(ck, list) else [ck],
        }]
        result["_review_needed"] = True
        result["_review_reason"] = "Multiple Answer - ověřit multi-select v UI"
    elif qtype == "Fill in the Blank Plus":
        # Toto potřebuje ruční konverzi. Zatím jen record.
        result["results"] = [{
            "key": "vysledek",
            "label_latex": "= ",
            "type": "mathlive",
            "expected": "; ".join(item.get("raw_correct_answers", [])) or "",
            "tolerance": 0,
        }]
        result["_review_needed"] = True
        result["_review_reason"] = "Fill in the Blank Plus - konvertovat na MC ručně"
    else:
        return None

    return result


# ------ Main ------------------------------------------------------------

def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--src", required=True, help="Adresář s rozbalenými fondy.")
    parser.add_argument("--out-json", required=True, help="Výstup JSONL (jeden řádek per úloha).")
    parser.add_argument("--out-xlsx", required=True, help="Výstup XLSX summary.")
    args = parser.parse_args()

    src = Path(args.src)
    if not src.is_dir():
        print(f"Chyba: {src} není adresář.", file=sys.stderr)
        return 2

    from collections import Counter
    all_tasks: list[dict[str, Any]] = []
    skipped: list[tuple[str, str, str]] = []  # (fond, qtype, reason)
    fond_counter: Counter = Counter()
    qtype_counter: Counter = Counter()

    for fond_dir in sorted(src.iterdir()):
        if not fond_dir.is_dir(): continue
        fond_name = fond_dir.name
        # Najdi hlavní res*.dat s items
        best_file, best_count = None, 0
        for dat in fond_dir.glob("res*.dat"):
            with open(dat, encoding='utf-8', errors='replace') as f:
                content = f.read()
            c = len(re.findall(r'<item\s', content))
            if c > best_count:
                best_count = c; best_file = dat
        if not best_file:
            print(f"⚠ Fond {fond_name}: nenalezen soubor s <item>", file=sys.stderr)
            continue
        with open(best_file, encoding='utf-8', errors='replace') as f:
            xml = f.read()
        items = extract_items_from_xml(xml)
        for idx, item_xml in enumerate(items, start=1):
            parsed = parse_item(item_xml)
            qtype_counter[parsed["question_type"] or "?"] += 1
            adapt = to_adaptmath(parsed, fond_name, idx)
            if adapt is None:
                skipped.append((fond_name, parsed["question_type"] or "?", "unsupported/broken"))
                continue
            all_tasks.append(adapt)
            fond_counter[fond_name] += 1

    print(f"Extrahováno {len(all_tasks)} úloh, přeskočeno {len(skipped)}.")
    print("\nRozdělení typů (všechny):")
    for t, c in qtype_counter.most_common():
        print(f"  {t}: {c}")

    # ------ JSON output -----------------------------------------------
    with open(args.out_json, "w", encoding='utf-8') as f:
        for t in all_tasks:
            f.write(json.dumps(t, ensure_ascii=False) + "\n")
    print(f"\nJSON preview: {args.out_json}")

    # ------ XLSX output ----------------------------------------------
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("openpyxl není nainstalovaný — XLSX přeskočeno.")
        return 0

    wb = Workbook()
    ws = wb.active
    ws.title = "Úlohy"
    headers = ["task_id", "fond", "typ_orig", "category", "content_latex",
               "expected/options", "review_needed", "review_reason"]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    for t in all_tasks:
        exp = ""
        if t["results"][0]["type"] == "multiple_choice":
            opts = t["results"][0]["options"]
            ck = t["results"][0]["expected"]
            exp = f"[MC {len(opts)}× → {ck}] " + " | ".join(
                f'{o["key"]}) {o["label_latex"][:60]}' for o in opts
            )
        else:
            exp = f"[{t['results'][0]['type']}] {t['results'][0]['expected']}"
        row = [
            t["task_id"],
            t["_source_fond"],
            t["_source_qtype"],
            t["category"] or "",
            t["content_latex"][:400],
            exp[:400],
            "YES" if t["_review_needed"] else "",
            t["_review_reason"],
        ]
        ws.append(row)

    for col in ws.columns:
        max_len = min(80, max(len(str(c.value or "")) for c in col))
        ws.column_dimensions[col[0].column_letter].width = max_len + 2
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = Alignment(vertical="top", wrap_text=True)

    # ---- druhý list: souhrn per fond ----
    ws2 = wb.create_sheet("Souhrn fondů")
    ws2.append(["Fond", "Úloh použitelných", "Typ (přehled)"])
    for c in ws2[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    for fond, count in sorted(fond_counter.items()):
        types = Counter(t["_source_qtype"] for t in all_tasks if t["_source_fond"] == fond)
        types_str = ", ".join(f"{t}={c}" for t, c in types.most_common())
        ws2.append([fond, count, types_str])
    ws2.column_dimensions['A'].width = 45
    ws2.column_dimensions['B'].width = 20
    ws2.column_dimensions['C'].width = 60

    # ---- třetí list: skipped ----
    ws3 = wb.create_sheet("Přeskočené")
    ws3.append(["Fond", "Typ", "Důvod"])
    for c in ws3[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    skipped_counter = Counter((f, q) for f, q, _ in skipped)
    for (fond, qtype), count in sorted(skipped_counter.items()):
        ws3.append([fond, qtype, f"{count} × unsupported (File Upload / Jumbled / Ordering / Short Response)"])
    ws3.column_dimensions['A'].width = 45
    ws3.column_dimensions['B'].width = 25
    ws3.column_dimensions['C'].width = 80

    wb.save(args.out_xlsx)
    print(f"XLSX summary: {args.out_xlsx}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
