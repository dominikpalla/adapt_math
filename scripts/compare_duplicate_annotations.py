"""
Side-by-side srovnání anotací pro duplicitní páry (z plan_delete_duplicates).

Pro každou skupinu duplicit vypíše VŠECHNY její úlohy vedle sebe:
  category, task_type, properties, skills, cognitive_load,
  irt_difficulty, irt_discrimination

Cíl: Dominik uvidí, která úloha v páru je LÉPE otagovaná, a rozhodne,
kterou v každém páru ponechat.

Použití:
    DATABASE_URL=... python scripts/compare_duplicate_annotations.py \\
        --out-xlsx /tmp/dup_annotation_compare.xlsx
"""
from __future__ import annotations
import argparse, os, re, sys, json
from pathlib import Path
from collections import defaultdict
from difflib import SequenceMatcher
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from database import init_db
from model import MathTask


def normalize_latex(s: str) -> str:
    if not s: return ""
    s = s.strip()
    reps = [
        (r'\$', ''), (r'\\limits', ''), (r'\\dfrac', r'\\frac'),
        (r'\\tg\b', r'\\tan'), (r'\\cotg\b', r'\\cot'),
        (r'\\arctg\b', r'\\arctan'), (r'\\arccotg\b', r'\\arccot'),
        (r'\\Rightarrow', r'\\Ra'),
        (r'\\left\s*\(', r'('), (r'\\right\s*\)', r')'),
        (r'\\left\s*\[', r'['), (r'\\right\s*\]', r']'),
        (r'\\left\s*\\\{', r'\\{'), (r'\\right\s*\\\}', r'\\}'),
        (r'\\langle', r'['), (r'\\rangle', r']'),
        (r'\\text\s*\{[^}]*\}', ''),
        (r'\\quad|\\qquad|\\,|\\;|\\:|\\!', ''), (r'~', ''),
    ]
    for old, new in reps:
        s = re.sub(old, new, s)
    return re.sub(r'\s+', '', s).lower()


def content_key(t):
    parts = [normalize_latex(t.content_latex or "")]
    if isinstance(t.results, list):
        for r in t.results:
            if not isinstance(r, dict): continue
            exp = r.get("expected")
            if isinstance(exp, (str, int, float)):
                parts.append(normalize_latex(str(exp)))
            elif isinstance(exp, list):
                for e in exp: parts.append(normalize_latex(str(e)))
    return "||".join(parts)


def annot_score(t):
    s = 0
    if t.category: s += 3
    if isinstance(t.properties, list) and t.properties: s += 2
    if isinstance(t.task_type, list) and t.task_type: s += 1
    if isinstance(t.skills, list) and t.skills: s += 2
    if isinstance(t.knowledge_vector, dict) and t.knowledge_vector: s += 5
    if t.irt_difficulty is not None: s += 1
    if t.cognitive_load: s += 1
    return s


def joinlist(x):
    if not x: return ""
    if isinstance(x, list):
        return " | ".join(str(v) for v in x)
    if isinstance(x, dict):
        return " | ".join(f"{k}={v}" for k, v in x.items() if v)
    return str(x)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--db-url", default=os.environ.get("DATABASE_URL"))
    parser.add_argument("--out-xlsx", default="/tmp/dup_annotation_compare.xlsx")
    args = parser.parse_args()
    if not args.db_url:
        print("Chyba: DATABASE_URL.", file=sys.stderr); return 2

    SessionLocal = init_db(args.db_url)
    sess = SessionLocal()
    try:
        tasks = sess.query(MathTask).all()
        by_key = defaultdict(list)
        for t in tasks:
            k = content_key(t)
            if k.strip("|"):
                by_key[k].append(t)
        groups = [g for g in by_key.values() if len(g) >= 2]
        # sort groups by keeper id pro čitelnost
        groups.sort(key=lambda g: sorted(t.task_id for t in g)[0])
        print(f"EXACT skupin: {len(groups)}, celkem úloh: {sum(len(g) for g in groups)}")

        # XLSX
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            print("openpyxl chybí."); return 0

        wb = Workbook()
        ws = wb.active
        ws.title = "Duplicity srovnání"
        headers = [
            "Skupina", "task_id", "content_latex (zkráceno)",
            "category", "properties", "task_type", "skills",
            "cognitive_load", "irt_diff", "irt_disc",
            "knowledge_vector (klíčů)", "SKÓRE", "NÁVRH",
        ]
        ws.append(headers)
        for c in ws[1]:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
            c.alignment = Alignment(vertical="top", wrap_text=True)

        keep_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        del_fill  = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        alt1 = PatternFill(start_color="FAFAFA", end_color="FAFAFA", fill_type="solid")

        row_idx = 2
        for gi, g in enumerate(groups, 1):
            # keeper = nejvyšší skóre; tie-break priority zdroje / nižší číslo
            scored = [(annot_score(t), t) for t in g]
            max_score = max(s for s, _ in scored)
            keepers = [t for s, t in scored if s == max_score]
            if len(keepers) > 1:
                # tie-break: cv > umat > e > ss, pak nižší číslo
                def prio(t):
                    src = re.match(r'([a-z]+)', t.task_id).group(1)
                    order = {"cv": 0, "umat": 1, "e": 2, "ss": 3}.get(src, 9)
                    n = int("".join(re.findall(r'\d+', t.task_id))[:6] or "999999")
                    return (order, n)
                keepers.sort(key=prio)
            keeper_id = keepers[0].task_id

            for t in sorted(g, key=lambda x: x.task_id):
                is_keep = (t.task_id == keeper_id)
                kv_count = len(t.knowledge_vector) if isinstance(t.knowledge_vector, dict) else 0
                row = [
                    gi, t.task_id,
                    (t.content_latex or "")[:120],
                    t.category or "",
                    joinlist(t.properties),
                    joinlist(t.task_type),
                    joinlist(t.skills),
                    t.cognitive_load or "",
                    t.irt_difficulty if t.irt_difficulty is not None else "",
                    t.irt_discrimination if t.irt_discrimination is not None else "",
                    kv_count,
                    annot_score(t),
                    "← PONECHAT" if is_keep else "SMAZAT",
                ]
                ws.append(row)
                for c in ws[row_idx]:
                    c.alignment = Alignment(vertical="top", wrap_text=True)
                    c.fill = keep_fill if is_keep else del_fill
                row_idx += 1
            # oddělovací prázdný řádek
            ws.append([])
            row_idx += 1

        widths = [8, 15, 45, 22, 30, 20, 30, 10, 8, 8, 12, 7, 12]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[chr(ord('A')+i-1)].width = w

        # freeze header
        ws.freeze_panes = "A2"

        wb.save(args.out_xlsx)
        print(f"XLSX: {args.out_xlsx}")
        # Souhrn statistik pro terminál
        both_annotated = sum(
            1 for g in groups if all(annot_score(t) > 0 for t in g)
        )
        one_side_empty = sum(
            1 for g in groups if any(annot_score(t) == 0 for t in g)
        )
        print(f"\nSouhrn:")
        print(f"  Skupin kde OBĚ úlohy anotované: {both_annotated}")
        print(f"  Skupin kde JEDNA úloha vůbec nemá anotace: {one_side_empty}")
        return 0
    finally:
        sess.close()


if __name__ == "__main__":
    sys.exit(main())
