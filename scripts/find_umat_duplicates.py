"""
Detekce duplicit v UMAT úlohách (matematici to při procházení tušili).

Strategie:
  1) Přesná duplicita: hash normalizovaného content_latex + expected.
     Normalizace: strip whitespace, sjednotit \\frac vs \\dfrac, \\tan vs \\tg
     (protože oba již máme jako sémanticky ekvivalentní), lowercase LaTeX
     příkazy jako \\Frac -> \\frac, remove trailing punctuation.
  2) Fuzzy duplicita: pro úlohy bez exact match spočítat normalized
     Levenshtein similarity (pomocí difflib.SequenceMatcher) mezi všemi
     páry. Report páry s podobností >= 0.85.

Výstup: XLSX + terminal report.

Použití:
    DATABASE_URL=... python scripts/find_umat_duplicates.py \\
        --out-xlsx /tmp/umat_duplicates.xlsx
"""
from __future__ import annotations
import argparse, os, re, sys, hashlib, json
from pathlib import Path
from collections import defaultdict
from difflib import SequenceMatcher
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from database import init_db
from model import MathTask


def normalize_latex(s: str) -> str:
    """Sjednotí zápis pro srovnávání."""
    if not s:
        return ""
    s = s.strip()
    # Sjednotit synonyma
    replacements = [
        (r'\\dfrac', r'\\frac'),
        (r'\\tg\b', r'\\tan'),
        (r'\\cotg\b', r'\\cot'),
        (r'\\arctg\b', r'\\arctan'),
        (r'\\arccotg\b', r'\\arccot'),
        (r'\\Rightarrow', r'\\Ra'),
        (r'\\left\s*\(', r'('),
        (r'\\right\s*\)', r')'),
        (r'\\left\s*\[', r'['),
        (r'\\right\s*\]', r']'),
        (r'\\left\s*\\\{', r'\\{'),
        (r'\\right\s*\\\}', r'\\}'),
        (r'\\mathbb\{Z\}', r'\\Bbb Z'),
        (r'\\mathbf\{Z\}', r'\\Bbb Z'),
        (r'\\mathbf\s*Z', r'\\Bbb Z'),
        (r'\\{Z\}', r'Z'),
    ]
    for old, new in replacements:
        s = re.sub(old, new, s)
    # Whitespace normalize
    s = re.sub(r'\s+', ' ', s)
    # Remove non-content whitespace kolem separátorů
    s = re.sub(r'\s*([=+\-*/()\[\]{},.])\s*', r'\1', s)
    return s.lower()


def content_key(task) -> str:
    """Kanonický otisk pro exact match."""
    parts = [normalize_latex(task.content_latex or "")]
    if isinstance(task.results, list):
        for r in task.results:
            if not isinstance(r, dict): continue
            exp = r.get("expected")
            if isinstance(exp, (str, int, float)):
                parts.append(normalize_latex(str(exp)))
            elif isinstance(exp, list):
                for e in exp:
                    parts.append(normalize_latex(str(e)))
    return "||".join(parts)


def content_for_fuzzy(task) -> str:
    """Pro fuzzy jen content_latex (bez expected — matematici řeknou
    'to samé zadání' i když je odpověď jinak formátovaná)."""
    return normalize_latex(task.content_latex or "")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--db-url", default=os.environ.get("DATABASE_URL"))
    parser.add_argument("--out-xlsx", default="/tmp/umat_duplicates.xlsx")
    parser.add_argument("--fuzzy-threshold", type=float, default=0.85)
    parser.add_argument("--prefix", default="umat_", help="Task ID prefix pro filter")
    args = parser.parse_args()
    if not args.db_url:
        print("Chyba: DATABASE_URL.", file=sys.stderr); return 2

    SessionLocal = init_db(args.db_url)
    sess = SessionLocal()
    try:
        tasks = sess.query(MathTask).filter(
            MathTask.task_id.like(f"{args.prefix}%")
        ).all()
        print(f"Načteno {len(tasks)} úloh s prefixem {args.prefix!r}.")

        # 1) Exact groups
        by_key = defaultdict(list)
        for t in tasks:
            by_key[content_key(t)].append(t)
        exact_dupes = [g for g in by_key.values() if len(g) > 1]
        print(f"\nExact-match skupin: {len(exact_dupes)} (celkem {sum(len(g) for g in exact_dupes)} úloh)")
        for g in exact_dupes:
            ids = [t.task_id for t in g]
            print(f"  {ids} — {g[0].content_latex[:80]}...")

        # 2) Fuzzy match (jen mezi non-exact-duplicates)
        exact_ids = {t.task_id for g in exact_dupes for t in g}
        candidates = [t for t in tasks if t.task_id not in exact_ids]
        # Grupovat po prefixu (umat_06_, umat_07_...) abychom nedělali cross-topic
        # matching (limity vs derivace nedávají smysl).
        groups = defaultdict(list)
        for t in candidates:
            # umat_XX_NN -> prefix "umat_XX"
            m = re.match(r'(umat_\d+)_', t.task_id)
            groups[m.group(1) if m else "other"].append(t)

        fuzzy_pairs = []
        for gname, gtasks in groups.items():
            if len(gtasks) < 2: continue
            n = len(gtasks)
            for i in range(n):
                ai = content_for_fuzzy(gtasks[i])
                if len(ai) < 15: continue  # příliš krátké → false positives
                for j in range(i+1, n):
                    aj = content_for_fuzzy(gtasks[j])
                    if len(aj) < 15: continue
                    # quick length filter: pokud délky lišší o >30%, skip
                    lr = min(len(ai), len(aj)) / max(len(ai), len(aj))
                    if lr < 0.6: continue
                    sim = SequenceMatcher(None, ai, aj).ratio()
                    if sim >= args.fuzzy_threshold:
                        fuzzy_pairs.append((gtasks[i], gtasks[j], sim))

        fuzzy_pairs.sort(key=lambda x: -x[2])
        print(f"\nFuzzy-similar párů (sim >= {args.fuzzy_threshold}): {len(fuzzy_pairs)}")
        for a, b, sim in fuzzy_pairs[:20]:
            print(f"  {sim:.2f}: {a.task_id} vs {b.task_id}")
            print(f"    A: {a.content_latex[:100]}")
            print(f"    B: {b.content_latex[:100]}")

        # ---- XLSX out ----
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            print("openpyxl chybí, XLSX přeskočen."); return 0

        wb = Workbook()
        ws = wb.active
        ws.title = "Exact duplicates"
        ws.append(["Skupina", "task_id", "content_latex", "expected (první)"])
        for c in ws[1]:
            c.font = Font(bold=True); c.fill = PatternFill(start_color="FFDDDD", end_color="FFDDDD", fill_type="solid")
        for idx, g in enumerate(exact_dupes, 1):
            for t in g:
                exp = ""
                if isinstance(t.results, list) and t.results:
                    r0 = t.results[0]
                    if isinstance(r0, dict):
                        exp = str(r0.get("expected", ""))[:200]
                ws.append([idx, t.task_id, (t.content_latex or "")[:400], exp])
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 70
        ws.column_dimensions['D'].width = 40

        ws2 = wb.create_sheet("Fuzzy similar")
        ws2.append(["Podobnost", "task_id A", "content A", "task_id B", "content B"])
        for c in ws2[1]:
            c.font = Font(bold=True); c.fill = PatternFill(start_color="FFF3DD", end_color="FFF3DD", fill_type="solid")
        for a, b, sim in fuzzy_pairs:
            ws2.append([
                round(sim, 3),
                a.task_id, (a.content_latex or "")[:400],
                b.task_id, (b.content_latex or "")[:400],
            ])
        ws2.column_dimensions['A'].width = 10
        ws2.column_dimensions['B'].width = 15
        ws2.column_dimensions['C'].width = 55
        ws2.column_dimensions['D'].width = 15
        ws2.column_dimensions['E'].width = 55
        for row in ws2.iter_rows(min_row=2):
            for c in row:
                c.alignment = Alignment(vertical="top", wrap_text=True)

        wb.save(args.out_xlsx)
        print(f"\nXLSX: {args.out_xlsx}")
        return 0
    finally:
        sess.close()


if __name__ == "__main__":
    sys.exit(main())
