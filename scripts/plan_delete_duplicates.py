"""
Plán mazání duplicit v celé DB.

Cíl: identifikovat OPRAVDU shodné úlohy (jen s jiným formátováním, ne
stylistické série se stejným typem, jinými hodnotami) a navrhnout,
kterou v každé skupině ponechat a které smazat.

Detekce:
  - Exact match: hash agresivně normalizovaného content_latex + expected
  - Fuzzy match: SequenceMatcher ratio >= 0.97 (jen velmi podobné,
    aby stylistické série s jinými hodnotami se nebraly)

Priorita pro PONECHÁNÍ:
  1) cv (ručně přidané, mají anotace) > umat > e > ss > olz*
  2) v rámci zdroje: úloha s nižším číslem (starší, více odkazů)
  3) tie-breaker: úloha s lépe vyplněnou anotací (category, properties, ...)

Výstup XLSX + terminal. Nic nemaže.

Použití:
    DATABASE_URL=... python scripts/plan_delete_duplicates.py \\
        --out-xlsx /tmp/delete_plan.xlsx
"""
from __future__ import annotations
import argparse, os, re, sys, json
from pathlib import Path
from collections import defaultdict
from difflib import SequenceMatcher
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from database import init_db
from model import MathTask


def normalize_latex(s: str) -> str:
    """Agresivní normalizace — odstraní vše, co může být jen formátovací rozdíl."""
    if not s:
        return ""
    s = s.strip()
    replacements = [
        (r'\$', ''),
        (r'\\limits', ''),
        (r'\\dfrac', r'\\frac'),
        (r'\\tg\b', r'\\tan'),
        (r'\\cotg\b', r'\\cot'),
        (r'\\arctg\b', r'\\arctan'),
        (r'\\arccotg\b', r'\\arccot'),
        (r'\\Rightarrow', r'\\Ra'),
        (r'\\left\s*\(', r'('),
        (r'\\right\s*\)', r')'),
        (r'\\left\s*\[', r'['),
        (r'\\right\s*\]', r']'),
        (r'\\left\s*\\\{', r'\\{'),
        (r'\\right\s*\\\}', r'\\}'),
        (r'\\langle', r'['),
        (r'\\rangle', r']'),
        (r'\\text\s*\{[^}]*\}', ''),  # odstranit \text{...}
        (r'\\quad|\\qquad|\\,|\\;|\\:|\\!', ''),  # spacing macros
        (r'~', ''),
    ]
    for old, new in replacements:
        s = re.sub(old, new, s)
    s = re.sub(r'\s+', '', s)
    return s.lower()


def source_prefix(task_id: str) -> str:
    """cv/umat/e/ss/olz1/olz2/olzk/olte/..."""
    m = re.match(r'([a-z]+\d*[a-z]*)_', task_id)
    return m.group(1) if m else task_id.split('_')[0] if '_' in task_id else task_id


def content_key(task) -> str:
    parts = [normalize_latex(task.content_latex or "")]
    if isinstance(task.results, list):
        for r in task.results:
            if not isinstance(r, dict): continue
            exp = r.get("expected")
            if isinstance(exp, (str, int, float)):
                parts.append(normalize_latex(str(exp)))
            elif isinstance(exp, list):
                for e in exp:
                    parts.append(normalize_latex(str(e)))
    return "||".join(parts)


def content_for_fuzzy(task) -> str:
    parts = [normalize_latex(task.content_latex or "")]
    if isinstance(task.results, list):
        for r in task.results:
            if not isinstance(r, dict): continue
            exp = r.get("expected")
            if isinstance(exp, (str, int, float)):
                parts.append(normalize_latex(str(exp)))
    return "||".join(parts)


SOURCE_PRIORITY = {
    "cv": 100,
    "umat": 80,
    "e": 60,
    "ss": 40,
    # olz*, olte, oliva* — nízká priorita (dosud nejsou v DB)
}


def annotation_score(task) -> int:
    """Vyšší = lépe anotovaná."""
    score = 0
    if task.category: score += 3
    if isinstance(task.properties, list) and task.properties: score += 2
    if isinstance(task.task_type, list) and task.task_type: score += 1
    if isinstance(task.skills, list) and task.skills: score += 2
    if isinstance(task.knowledge_vector, dict) and task.knowledge_vector: score += 5
    return score


def task_id_num(task_id: str) -> int:
    """Extrahuje číslo z konce task_id (např. 'umat_08_39' → 39, 'cv05_05' → 5, 'e03_35_9' → 359)."""
    nums = re.findall(r'\d+', task_id)
    if not nums: return 999999
    return int("".join(nums[-2:] if len(nums) >= 2 else nums))


def pick_keeper(tasks):
    """Vyber úlohu k ponechání ze skupiny."""
    def key(t):
        src = source_prefix(t.task_id)
        base_src = re.sub(r'\d+$', '', src)  # cv05 → cv, umat → umat
        # (higher is better → negate for ascending sort)
        return (
            -SOURCE_PRIORITY.get(base_src, 20),  # vyšší priorita první
            -annotation_score(t),                 # více anotací první
            task_id_num(t.task_id),               # menší číslo první
            t.task_id,                            # tie-break lexikograficky
        )
    return sorted(tasks, key=key)[0]


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--db-url", default=os.environ.get("DATABASE_URL"))
    parser.add_argument("--out-xlsx", default="/tmp/delete_plan.xlsx")
    parser.add_argument("--fuzzy-threshold", type=float, default=0.97)
    args = parser.parse_args()
    if not args.db_url:
        print("Chyba: DATABASE_URL.", file=sys.stderr); return 2

    SessionLocal = init_db(args.db_url)
    sess = SessionLocal()
    try:
        tasks = sess.query(MathTask).all()
        print(f"Načteno {len(tasks)} úloh celkem.")

        # 1) Exact groups
        by_key = defaultdict(list)
        for t in tasks:
            k = content_key(t)
            if k.strip("|"):
                by_key[k].append(t)
        exact_groups = [g for g in by_key.values() if len(g) >= 2]
        exact_ids = {t.task_id for g in exact_groups for t in g}
        print(f"\nEXACT skupin (identický normalizovaný obsah): {len(exact_groups)}")
        print(f"  Celkem úloh v exact skupinách: {sum(len(g) for g in exact_groups)}")

        # 2) Fuzzy match >= 0.97 (jen VELMI podobné, jinak stylistické série)
        candidates = [t for t in tasks if t.task_id not in exact_ids and len(content_for_fuzzy(t)) >= 15]
        norm_map = [(t, content_for_fuzzy(t)) for t in candidates]
        n = len(norm_map)
        print(f"\nFuzzy porovnání {n} kandidátů (bez exact) při threshold >= {args.fuzzy_threshold}...")
        fuzzy_pairs = []
        for i in range(n):
            ta, ai = norm_map[i]
            for j in range(i+1, n):
                tb, aj = norm_map[j]
                lr = min(len(ai), len(aj)) / max(len(ai), len(aj))
                if lr < 0.85: continue
                sim = SequenceMatcher(None, ai, aj).ratio()
                if sim >= args.fuzzy_threshold:
                    fuzzy_pairs.append((ta, tb, sim))
        print(f"Fuzzy párů: {len(fuzzy_pairs)}")

        # Sjednotit fuzzy páry do skupin (transitive closure)
        parent = {}
        def find(x):
            while parent.get(x, x) != x:
                parent[x] = parent.get(parent[x], parent[x])
                x = parent[x]
            return x
        def union(a, b):
            ra, rb = find(a), find(b)
            if ra != rb: parent[ra] = rb

        fuzzy_task_map = {}
        for a, b, _ in fuzzy_pairs:
            fuzzy_task_map[a.task_id] = a
            fuzzy_task_map[b.task_id] = b
            union(a.task_id, b.task_id)
        fuzzy_groups_dict = defaultdict(list)
        for tid, t in fuzzy_task_map.items():
            fuzzy_groups_dict[find(tid)].append(t)
        fuzzy_groups = [g for g in fuzzy_groups_dict.values() if len(g) >= 2]
        print(f"Fuzzy skupin: {len(fuzzy_groups)}")

        # 3) Pro každou skupinu vybrat keeper
        delete_plan = []
        for g in exact_groups + fuzzy_groups:
            keeper = pick_keeper(g)
            for t in g:
                if t.task_id == keeper.task_id: continue
                delete_plan.append({
                    "match_type": "EXACT" if g in exact_groups else "FUZZY",
                    "group_size": len(g),
                    "delete_id": t.task_id,
                    "delete_content": (t.content_latex or "")[:200],
                    "delete_expected": _expected_str(t)[:100],
                    "keep_id": keeper.task_id,
                    "keep_content": (keeper.content_latex or "")[:200],
                    "keep_annot_score": annotation_score(keeper),
                    "delete_annot_score": annotation_score(t),
                })

        print(f"\n=== DELETE PLAN: {len(delete_plan)} úloh ke smazání ===")
        for d in delete_plan[:20]:
            print(f"  [{d['match_type']}] SMAZAT {d['delete_id']}  ← PONECHAT {d['keep_id']}")

        # ---- XLSX ----
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            print("openpyxl chybí, XLSX přeskočen."); return 0

        wb = Workbook()
        ws = wb.active
        ws.title = "Delete plan"
        headers = ["Match", "Skupina", "SMAZAT (task_id)", "SMAZAT: content",
                   "SMAZAT: expected", "PONECHAT (task_id)", "PONECHAT: content",
                   "Anotace keep", "Anotace del"]
        ws.append(headers)
        for c in ws[1]:
            c.font = Font(bold=True); c.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

        # sort by keeper_id pro čitelnost
        delete_plan.sort(key=lambda d: (d["match_type"], d["keep_id"], d["delete_id"]))
        for i, d in enumerate(delete_plan, 1):
            ws.append([
                d["match_type"],
                d["group_size"],
                d["delete_id"],
                d["delete_content"],
                d["delete_expected"],
                d["keep_id"],
                d["keep_content"],
                d["keep_annot_score"],
                d["delete_annot_score"],
            ])

        widths = [8, 8, 18, 50, 30, 18, 50, 12, 12]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[chr(ord('A')+i-1)].width = w
        for row in ws.iter_rows(min_row=2):
            for c in row:
                c.alignment = Alignment(vertical="top", wrap_text=True)

        # Druhý list: seznam pro smazání (jen task_ids)
        ws2 = wb.create_sheet("SQL delete list")
        ws2.append(["task_id ke smazání"])
        ws2['A1'].font = Font(bold=True)
        for d in delete_plan:
            ws2.append([d["delete_id"]])
        ws2.column_dimensions['A'].width = 25

        wb.save(args.out_xlsx)
        print(f"\nXLSX: {args.out_xlsx}")
        print(f"Souhrn: {len(exact_groups)} exact + {len(fuzzy_groups)} fuzzy skupin → {len(delete_plan)} úloh ke smazání")
        return 0
    finally:
        sess.close()


def _expected_str(task) -> str:
    if not isinstance(task.results, list) or not task.results: return ""
    r = task.results[0]
    if not isinstance(r, dict): return ""
    return str(r.get("expected", ""))


if __name__ == "__main__":
    sys.exit(main())
