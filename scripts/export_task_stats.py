"""
Export statistik anotace úloh do XLSX pro dr. Medkovou.

Sestaví komplexní Excel s několika listy:

  1) Souhrn             — hlavní čísla (total, anotováno, po cvičeních).
  2) Kategorie          — histogram: kategorie × počet úloh + % z anotovaných.
  3) Vlastnosti         — histogram: vlastnost × počet úloh (multi-select).
  4) Typy               — histogram: typ × počet úloh.
  5) Dovednosti         — histogram: dovednost × počet úloh.
  6) Kategorie × Vlastnosti — cross-tab (řádky = kategorie, sloupce = vlastnosti).
  7) Kategorie × Dovednosti — cross-tab.
  8) Seznam úloh        — plný seznam všech 415 úloh s anotací (jedna řada).
  9) Neanotované úlohy  — úlohy, které nemají vyplněnou žádnou anotaci.

Použití:
    DATABASE_URL=postgresql://... python scripts/export_task_stats.py [--out FILE]

Beze zásahu do DB (READ-ONLY).
"""

from __future__ import annotations

import argparse
import os
import sys
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

# Přidat parent dir do sys.path, ať fungují importy jako `from tasks...`
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from model import MathTask
from tasks.knowledge_weights import (
    TASK_CATEGORIES, TASK_PROPERTIES, TASK_TYPES, TASK_SKILLS,
    WEIGHT_GROUPS, GROUP_LABELS,
)


# --- Styling ----------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="F37021")   # AdaptMath oranžová
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

SUBHEADER_FILL = PatternFill("solid", fgColor="FFE4C2")
SUBHEADER_FONT = Font(bold=True, color="8B4513", size=10)

TOTAL_ROW_FILL = PatternFill("solid", fgColor="FFF3E0")
TOTAL_ROW_FONT = Font(bold=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="D0D5D9"),
    right=Side(style="thin", color="D0D5D9"),
    top=Side(style="thin", color="D0D5D9"),
    bottom=Side(style="thin", color="D0D5D9"),
)

# Barvy pro group_key (odpovídají CSS třídám .grp-*)
GROUP_COLORS = {
    "vlasnosti":  "E8F5E9",  # světle zelená
    "typ":        "FFE0B2",
    "dovednosti": "FFCCBC",
    "ss":         "FFF9C4",
    "logika":     "CFD8DC",
    "funkce":     "BBDEFB",
    "monotonie":  "E1BEE7",
    "konvex":     "CFD8DC",
    "spojitost":  "B2DFDB",
    "limita":     "F8BBD0",
    "derivace":   "C5CAE9",
    "prubeh":     "EEEEEE",
    "pf":         "FFCDD2",
    "ui":         "D7CCC8",
}


def style_header(cell) -> None:
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = HEADER_ALIGN
    cell.border = THIN_BORDER


def style_cell(cell, bold: bool = False, align: str = "left", bg: str | None = None) -> None:
    if bold:
        cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal=align, vertical="top", wrap_text=True)
    cell.border = THIN_BORDER
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)


def autosize(ws, widths: dict[int, int]) -> None:
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


# --- Data loading -----------------------------------------------------------

def load_tasks(db_url: str) -> list[dict]:
    engine = create_engine(db_url)
    SessionLocal = sessionmaker(bind=engine)
    session = SessionLocal()
    try:
        rows = session.query(MathTask).order_by(MathTask.task_id).all()
        out = []
        for t in rows:
            out.append({
                "task_id": t.task_id,
                "cognitive_load": t.cognitive_load,
                "category": t.category,
                "properties": t.properties or [],
                "task_type": t.task_type or [],
                "skills": t.skills or [],
                "content_latex": t.content_latex,
                "results_types": [r.get("type") for r in (t.results or [])],
            })
        return out
    finally:
        session.close()
        engine.dispose()


def exercise_of(task_id: str) -> str:
    """`cv04_15` → `cv04`, `cv12_1` → `cv12`, `` → `??`."""
    if not task_id or "_" not in task_id:
        return "??"
    return task_id.split("_", 1)[0]


def is_annotated(task: dict) -> bool:
    """Úloha se považuje za anotovanou, pokud má vyplněné cokoli
    z 4 hlavních anotačních polí."""
    return bool(
        task["category"]
        or task["properties"]
        or task["task_type"]
        or task["skills"]
    )


# --- Sheet builders ---------------------------------------------------------

def sheet_souhrn(wb: Workbook, tasks: list[dict]) -> None:
    ws = wb.create_sheet("Souhrn", 0)

    total = len(tasks)
    n_annot = sum(1 for t in tasks if is_annotated(t))
    n_cat = sum(1 for t in tasks if t["category"])
    n_prop = sum(1 for t in tasks if t["properties"])
    n_type = sum(1 for t in tasks if t["task_type"])
    n_skill = sum(1 for t in tasks if t["skills"])
    n_cog = sum(1 for t in tasks if t["cognitive_load"])

    ws["A1"] = "AdaptMath — statistika anotace úloh"
    ws["A1"].font = Font(bold=True, size=14, color="F37021")
    ws.merge_cells("A1:B1")

    ws["A2"] = f"Vygenerováno: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = Font(italic=True, size=9, color="888888")
    ws.merge_cells("A2:B2")

    rows = [
        ("Celkem úloh v DB", total),
        ("Anotováno (min. 1 z 4 polí vyplněné)", n_annot),
        ("Neanotováno", total - n_annot),
        ("% anotováno", f"{n_annot / total * 100:.1f} %"),
        ("", ""),
        ("Úloh s kategorií", n_cat),
        ("Úloh s vlastnostmi", n_prop),
        ("Úloh s typem", n_type),
        ("Úloh s dovednostmi", n_skill),
        ("Úloh s cognitive_load", n_cog),
    ]
    for i, (label, val) in enumerate(rows, start=4):
        c1, c2 = ws.cell(row=i, column=1, value=label), ws.cell(row=i, column=2, value=val)
        if label:
            c1.font = Font(bold=True)
        c1.alignment = Alignment(horizontal="left", vertical="center")
        c2.alignment = Alignment(horizontal="right", vertical="center")

    # Po cvičeních
    start = len(rows) + 6
    ws.cell(row=start, column=1, value="Rozdělení po cvičeních")
    ws.cell(row=start, column=1).font = Font(bold=True, size=12, color="F37021")
    ws.merge_cells(start_row=start, start_column=1, end_row=start, end_column=4)

    hdrs = ["Cvičení", "Celkem úloh", "Anotováno", "%"]
    for j, h in enumerate(hdrs, start=1):
        c = ws.cell(row=start + 1, column=j, value=h)
        style_header(c)

    by_ex = defaultdict(lambda: [0, 0])   # [total, annotated]
    for t in tasks:
        ex = exercise_of(t["task_id"])
        by_ex[ex][0] += 1
        if is_annotated(t):
            by_ex[ex][1] += 1

    r = start + 2
    for ex in sorted(by_ex):
        tot, ann = by_ex[ex]
        vals = [ex, tot, ann, f"{ann / tot * 100:.0f} %" if tot else "-"]
        for j, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=j, value=v)
            style_cell(c, align="left" if j == 1 else "right", bold=(j == 1))
        r += 1

    # Total row
    tot_all = sum(v[0] for v in by_ex.values())
    ann_all = sum(v[1] for v in by_ex.values())
    for j, v in enumerate(["CELKEM", tot_all, ann_all, f"{ann_all / tot_all * 100:.1f} %"], start=1):
        c = ws.cell(row=r, column=j, value=v)
        style_cell(c, bold=True, align="left" if j == 1 else "right", bg="FFF3E0")

    autosize(ws, {1: 45, 2: 15, 3: 15, 4: 10})


def sheet_histogram(
    wb: Workbook, tasks: list[dict], title: str,
    all_values: list[str], task_field: str,
    is_multi: bool, position: int,
) -> None:
    """Univerzální histogram: pro každou hodnotu z `all_values` spočítej,
    kolik úloh ji má. Pro multi-select (properties/task_type/skills)
    počítáme výskyty v listu; pro single-value (category) jen matches."""
    ws = wb.create_sheet(title, position)

    # Nakreslíme počty
    counts = Counter()
    if is_multi:
        for t in tasks:
            for v in t[task_field] or []:
                counts[v] += 1
    else:
        for t in tasks:
            if t[task_field]:
                counts[t[task_field]] += 1

    # Skupinové součty (podle WEIGHT_GROUPS)
    group_totals = defaultdict(int)
    for v, n in counts.items():
        group_totals[WEIGHT_GROUPS.get(v, "default")] += n

    # Header
    hdrs = ["Skupina", "Hodnota", "Počet úloh", "% z celku"]
    for j, h in enumerate(hdrs, start=1):
        c = ws.cell(row=1, column=j, value=h)
        style_header(c)

    total_annot = sum(1 for t in tasks if is_annotated(t))
    r = 2

    # Řadit dle původního pořadí z KNOWLEDGE_WEIGHTS + přidat i položky
    # s nula výskyty (přehled kompletnosti). Sešup podle skupiny.
    for v in all_values:
        n = counts.get(v, 0)
        pct = f"{n / total_annot * 100:.1f} %" if total_annot else "-"
        grp = WEIGHT_GROUPS.get(v, "default")
        grp_label = GROUP_LABELS.get(grp, grp)
        vals = [grp_label, v, n, pct]
        bg = GROUP_COLORS.get(grp)
        for j, val in enumerate(vals, start=1):
            c = ws.cell(row=r, column=j, value=val)
            style_cell(c, align="left" if j <= 2 else "right", bg=bg)
        # Pokud n == 0, mírně ztlumíme
        if n == 0:
            for j in range(1, 5):
                c = ws.cell(row=r, column=j)
                c.font = Font(color="AAAAAA", italic=True)
        r += 1

    # Total
    tot = sum(counts.values())
    for j, v in enumerate(["", "CELKEM VÝSKYTŮ", tot,
                            f"{tot / total_annot * 100:.1f} %" if total_annot else "-"], start=1):
        c = ws.cell(row=r, column=j, value=v)
        style_cell(c, bold=True, align="left" if j <= 2 else "right", bg="FFF3E0")

    ws.freeze_panes = "A2"
    autosize(ws, {1: 22, 2: 45, 3: 13, 4: 12})


def sheet_crosstab(
    wb: Workbook, tasks: list[dict], title: str,
    row_field: str, col_field: str, col_values: list[str],
    row_is_multi: bool, col_is_multi: bool, position: int,
) -> None:
    """Kontingenční tabulka: řádky = unikátní hodnoty z row_field, sloupce = col_values.
    Buňka [řádek, sloupec] = počet úloh, které mají obě hodnoty současně."""
    ws = wb.create_sheet(title, position)

    row_values_set = set()
    for t in tasks:
        rv = t[row_field]
        if not rv:
            continue
        if row_is_multi:
            row_values_set.update(rv)
        else:
            row_values_set.add(rv)
    # Řadit dle původního pořadí (kategorie mají významné pořadí)
    row_values = [v for v in TASK_CATEGORIES if v in row_values_set]
    # Přidat i cokoli, co v TASK_CATEGORIES není (šlo by o legacy)
    for rv in sorted(row_values_set - set(row_values)):
        row_values.append(rv)

    # Header
    hdrs = ["Kategorie \\ " + title.split(" × ")[-1]] + col_values + ["ŘÁDEK Σ"]
    for j, h in enumerate(hdrs, start=1):
        c = ws.cell(row=1, column=j, value=h)
        style_header(c)

    col_totals = [0] * len(col_values)
    r = 2
    for rv in row_values:
        # Row cell
        grp = WEIGHT_GROUPS.get(rv, "default")
        c = ws.cell(row=r, column=1, value=rv)
        style_cell(c, bold=True, bg=GROUP_COLORS.get(grp))

        row_total = 0
        for jc, cv in enumerate(col_values):
            # Kolik úloh má tuto řádkovou hodnotu i tento sloupcový tag?
            cnt = 0
            for t in tasks:
                # Match řádek
                if row_is_multi:
                    if rv not in (t[row_field] or []):
                        continue
                else:
                    if t[row_field] != rv:
                        continue
                # Match sloupec
                if col_is_multi:
                    if cv in (t[col_field] or []):
                        cnt += 1
                else:
                    if t[col_field] == cv:
                        cnt += 1
            cell = ws.cell(row=r, column=jc + 2, value=cnt if cnt else "")
            style_cell(cell, align="right",
                       bg=(GROUP_COLORS.get(WEIGHT_GROUPS.get(cv, "default")) if cnt else None))
            col_totals[jc] += cnt
            row_total += cnt
        # Row total
        cell = ws.cell(row=r, column=len(col_values) + 2, value=row_total)
        style_cell(cell, bold=True, align="right", bg="FFF3E0")
        r += 1

    # Column totals row
    ws.cell(row=r, column=1, value="SLOUPEC Σ")
    style_cell(ws.cell(row=r, column=1), bold=True, bg="FFF3E0")
    for jc, tot in enumerate(col_totals):
        cell = ws.cell(row=r, column=jc + 2, value=tot if tot else "")
        style_cell(cell, bold=True, align="right", bg="FFF3E0")
    # Grand total
    ws.cell(row=r, column=len(col_values) + 2, value=sum(col_totals))
    style_cell(ws.cell(row=r, column=len(col_values) + 2), bold=True, align="right", bg="FFF3E0")

    ws.freeze_panes = "B2"
    widths = {1: 35}
    for j in range(2, len(col_values) + 2):
        widths[j] = 14
    widths[len(col_values) + 2] = 12
    autosize(ws, widths)


def sheet_all_tasks(wb: Workbook, tasks: list[dict], position: int) -> None:
    ws = wb.create_sheet("Seznam úloh", position)

    hdrs = ["Task ID", "Cvičení", "Kognit. zátěž", "Kategorie",
            "Vlastnosti", "Typy", "Dovednosti", "Typy výsledků", "Zadání (LaTeX)"]
    for j, h in enumerate(hdrs, start=1):
        c = ws.cell(row=1, column=j, value=h)
        style_header(c)

    for i, t in enumerate(tasks, start=2):
        vals = [
            t["task_id"] or "(prázdné)",
            exercise_of(t["task_id"]),
            t["cognitive_load"] or "",
            t["category"] or "",
            ", ".join(t["properties"] or []),
            ", ".join(t["task_type"] or []),
            ", ".join(t["skills"] or []),
            ", ".join(t["results_types"] or []),
            (t["content_latex"] or "")[:200],
        ]
        for j, v in enumerate(vals, start=1):
            c = ws.cell(row=i, column=j, value=v)
            grp_bg = None
            if j == 4 and t["category"]:
                grp_bg = GROUP_COLORS.get(WEIGHT_GROUPS.get(t["category"], "default"))
            style_cell(c, bg=grp_bg,
                       align="left" if j != 3 else "center",
                       bold=(j == 1))

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = ws.dimensions
    autosize(ws, {1: 12, 2: 8, 3: 11, 4: 32, 5: 40, 6: 24, 7: 32, 8: 22, 9: 60})


def sheet_unannotated(wb: Workbook, tasks: list[dict], position: int) -> None:
    ws = wb.create_sheet("Neanotované", position)

    unann = [t for t in tasks if not is_annotated(t)]

    ws["A1"] = f"Úloh bez jakékoli anotace: {len(unann)} z {len(tasks)}"
    ws["A1"].font = Font(bold=True, size=12, color="F37021")
    ws.merge_cells("A1:C1")

    hdrs = ["Task ID", "Cvičení", "Zadání (LaTeX)"]
    for j, h in enumerate(hdrs, start=1):
        c = ws.cell(row=3, column=j, value=h)
        style_header(c)

    if not unann:
        c = ws.cell(row=4, column=1, value="✅ Všechny úlohy jsou anotované!")
        c.font = Font(bold=True, color="1E7E34", italic=True)
        ws.merge_cells("A4:C4")
    else:
        for i, t in enumerate(unann, start=4):
            vals = [t["task_id"] or "(prázdné)",
                    exercise_of(t["task_id"]),
                    (t["content_latex"] or "")[:200]]
            for j, v in enumerate(vals, start=1):
                c = ws.cell(row=i, column=j, value=v)
                style_cell(c, bold=(j == 1))

    autosize(ws, {1: 12, 2: 8, 3: 90})


# --- Main -------------------------------------------------------------------

def build_workbook(tasks: list[dict], out_path: Path) -> None:
    wb = Workbook()
    # Odstranit default "Sheet" — přidáme si vlastní
    default = wb.active
    wb.remove(default)

    sheet_souhrn(wb, tasks)
    sheet_histogram(wb, tasks, "Kategorie", TASK_CATEGORIES, "category", is_multi=False, position=1)
    sheet_histogram(wb, tasks, "Vlastnosti", TASK_PROPERTIES, "properties", is_multi=True, position=2)
    sheet_histogram(wb, tasks, "Typy", TASK_TYPES, "task_type", is_multi=True, position=3)
    sheet_histogram(wb, tasks, "Dovednosti", TASK_SKILLS, "skills", is_multi=True, position=4)
    sheet_crosstab(wb, tasks, "Kategorie × Vlastnosti", "category", "properties",
                   TASK_PROPERTIES, row_is_multi=False, col_is_multi=True, position=5)
    sheet_crosstab(wb, tasks, "Kategorie × Dovednosti", "category", "skills",
                   TASK_SKILLS, row_is_multi=False, col_is_multi=True, position=6)
    sheet_all_tasks(wb, tasks, position=7)
    sheet_unannotated(wb, tasks, position=8)

    wb.save(out_path)


def main() -> int:
    parser = argparse.ArgumentParser(description="Export AdaptMath tag statistics to XLSX.")
    parser.add_argument("--out", type=Path, default=Path("adaptmath_stats.xlsx"))
    parser.add_argument(
        "--db-url",
        default=os.environ.get("DATABASE_URL"),
        help="PostgreSQL URL. Fallback na env DATABASE_URL.",
    )
    args = parser.parse_args()

    if not args.db_url:
        print("Chyba: nastav DATABASE_URL nebo --db-url.", file=sys.stderr)
        return 2

    print(f"Načítám úlohy z DB...")
    tasks = load_tasks(args.db_url)
    print(f"  Úloh: {len(tasks)}")
    print(f"  Anotovaných: {sum(1 for t in tasks if is_annotated(t))}")

    print(f"Zapisuju do {args.out}...")
    args.out.parent.mkdir(parents=True, exist_ok=True)
    build_workbook(tasks, args.out)
    size_kb = args.out.stat().st_size / 1024
    print(f"  Hotovo: {args.out} ({size_kb:.1f} KB)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
